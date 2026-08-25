#!/usr/bin/env python3
"""
Rebuilds WHOLE_PRODUCT_LIST_BO_PRODUCT_CALENDAR.xlsx as nine per-category sheets.

The workbook that shipped before this script was a cartesian template: one sheet, nine
blocks, each block 24 specification headings x 22 attribute slots = 528 rows per category
whether or not the category has such a property. A light bulb carried "Burst Pressure",
"Pourability" and "Drainage Flow"; six of the nine categories repeated the whole "Product
identity" column verbatim under "Manufactring Specifications", "Hygiene specifications"
and "Packaging". 1,302 cells held U+FFFD where an em dash had been destroyed by an
encoding round-trip. Nothing in it carried a value, so none of it could be shown.

What this produces instead: one sheet per category, one row per attribute that at least
one SKU in that category actually has a value for, with the value for each of the three
brands beside it and the provenance of each value. The sheet is therefore both the schema
and the data, and it is the file a human edits when a specification changes.

  Source of schema : services/pipeline/registry/{category}.json
  Source of values : services/pipeline/data/curated/{category}/{SKU}.json
  Source of groups : services/pipeline/registry/spec-groups.json
  Consumed by      : services/pipeline/src/registry/from-sheet.ts  (sheet -> registry -> DB)

This is a migration, not part of the build. It ran once to turn the template into the
workbook; after that the workbook is edited directly and from-sheet.ts is what reads it.
Re-running it regenerates the workbook from the curated fixtures and will discard any
edit made in Excel since.

Run:  python services/pipeline/tools/build_calendar_workbook.py
"""

from __future__ import annotations

import json
import math
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
REGISTRY = ROOT / "services" / "pipeline" / "registry"
CURATED = ROOT / "services" / "pipeline" / "data" / "curated"
OUT = ROOT / "WHOLE_PRODUCT_LIST_BO_PRODUCT_CALENDAR.xlsx"

# ── the placeholder families ────────────────────────────────────────────────────
# Two strings were written across 12,898 of the 15,080 attribute values (85.5%), both
# stamped provenance "verified" at confidence 0.98. They are not values and they never
# reach a page; they are dropped here so the sheet only carries things that are true.
PLACEHOLDERS = {
    "verified industry standard",
    "certified architectural standard",
    "not applicable",
    "n/a",
    "na",
    "tbd",
    "-",
    "",
}
# `standards_and_certifications_parameter_#1 .. #23` exist only to hold the second of those.
JUNK_KEY = re.compile(r"_parameter_#\d+$")

# A category's SKUs are its three brands. Ambuja shipped two cement products; the catalogue
# is nine categories x three brands, so the second Ambuja SKU is not part of it.
EXCLUDE_SKUS = {"CEM-AMB-KAWACH50"}

# Column A of the old workbook: the department a category sits under. The storefront nav
# groups the nine live categories by these, and the subcategory beneath is the category.
DEPARTMENT = {
    "epoxy": "Construction Chemicals",
    "fire-extinguishers": "Safety & Fire",
    "solar-panels": "Solar & Energy",
    "cctv": "CCTV & Security",
    "tiles": "Building Materials",
    "glass": "Building Materials",
    "total-stations": "Surveying Equipment",
    "bulbs": "Electrical Items",
    "cement": "Construction Materials",
}

# ── key specifications ──────────────────────────────────────────────────────────
# The rows a buyer reads before anything else, in the order they read them. The old
# registries carried eight per category chosen by a generic importance rank, which is how
# a bulb ended up showing "Operating life" but not luminous intensity. These are the
# properties an Indian site engineer actually specifies against.
KEY_SPECS = {
    "bulbs": [
        "wattage",
        "luminous_flux",
        "luminous_intensity",
        "colour_temperature",
        "colour_rendering_index",
        "base_type",
        "beam_angle",
        "operating_life",
    ],
    "cement": [
        "cement_type",
        "compressive_strength_28d",
        "fineness_blaine",
        "setting_time_initial",
        "is_standard",
        "net_weight_kg",
        "recommended_concrete_grade",
        "shelf_life_months",
    ],
    "glass": [
        "nominal_thickness_mm",
        "visible_light_transmission_pct",
        "solar_heat_gain_coefficient",
        "u_value_w_m2k",
        "sound_reduction_db",
        "impact_safety_class",
        "light_to_solar_gain",
        "max_pane_width_mm",
    ],
    "tiles": [
        "dim_w_mm",
        "water_absorption_pct",
        "pei_abrasion_class",
        "slip_resistance_class",
        "surface_finish",
        "coverage_per_box_sqft",
        "rectified",
        "modulus_of_rupture_mpa",
    ],
    "solar-panels": [
        "rated_power_wp",
        "module_efficiency_pct",
        "cell_technology",
        "voc_v",
        "isc_a",
        "fill_factor",
        "temp_coefficient_pmax",
        "performance_warranty_years",
    ],
    "cctv": [
        "resolution_mp",
        "image_sensor",
        "lens_focal_length_mm",
        "ir_range_m",
        "min_illumination_lux",
        "ip_rating",
        "horizontal_field_of_view_deg",
        "form_factor",
    ],
    "total-stations": [
        "angular_accuracy_arcsec",
        "distance_accuracy_prism",
        "reflectorless_range_m",
        "prism_range_m",
        "compensator_type",
        "telescope_magnification_x",
        "battery_operating_time_h",
        "ip_rating",
    ],
    "fire-extinguishers": [
        "extinguishing_agent",
        "capacity_kg",
        "suitable_fire_classes",
        "fire_rating",
        "discharge_time_s",
        "discharge_range_m",
        "working_pressure_bar",
        "is_standard",
    ],
    "epoxy": [
        "epoxy_type",
        "compressive_strength_mpa",
        "bond_strength_concrete_mpa",
        "pot_life_min",
        "full_cure_time_days",
        "application_thickness_mm",
        "pack_size_kg",
        "coverage",
    ],
}

# The three specs printed on a product card in a listing grid.
CARD_SPECS = {
    "bulbs": ["wattage", "luminous_flux", "colour_temperature"],
    "cement": ["cement_type", "compressive_strength_28d", "net_weight_kg"],
    "glass": ["nominal_thickness_mm", "visible_light_transmission_pct", "u_value_w_m2k"],
    "tiles": ["dim_w_mm", "surface_finish", "pei_abrasion_class"],
    "solar-panels": ["rated_power_wp", "module_efficiency_pct", "cell_technology"],
    "cctv": ["resolution_mp", "lens_focal_length_mm", "ir_range_m"],
    "total-stations": ["angular_accuracy_arcsec", "reflectorless_range_m", "telescope_magnification_x"],
    "fire-extinguishers": ["extinguishing_agent", "capacity_kg", "fire_rating"],
    "epoxy": ["epoxy_type", "compressive_strength_mpa", "pack_size_kg"],
}


# ── derived specifications ──────────────────────────────────────────────────────
# Three properties a datasheet states that no source in this catalogue carried, each one
# an exact function of numbers that are carried. They are computed rather than guessed and
# are written with provenance "derived" so a reader can tell them apart from a fetched
# figure. The formula travels with the value into the sheet's Notes column.
def _num(v):
    """First number in a value, or None. '220-240 V AC' -> 220.0, 806 -> 806.0."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if not isinstance(v, str):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", v.replace(",", ""))
    return float(m.group()) if m else None


def derive_luminous_intensity(vals: dict) -> tuple[float, str] | None:
    """
    Mean luminous intensity in candela: I = flux / solid angle, with the solid angle of a
    cone of apex angle theta being 2*pi*(1 - cos(theta/2)) steradians. This is the property
    a lamp is chosen on when it has to light a specific plane, and the reason a 200-degree
    806 lm lamp and a 120-degree 806 lm lamp are not interchangeable.
    """
    flux = _num(vals.get("luminous_flux")) or _num(vals.get("lumens"))
    beam = _num(vals.get("beam_angle"))
    if not flux or not beam or not (0 < beam <= 360):
        return None
    omega = 2 * math.pi * (1 - math.cos(math.radians(beam) / 2))
    if omega <= 0:
        return None
    return round(flux / omega, 1), "I = luminous flux / (2π(1 − cos(beam angle / 2))), the beam's solid angle in steradians"


def derive_light_to_solar_gain(vals: dict) -> tuple[float, str] | None:
    """
    LSG = visible light transmission / solar heat gain coefficient. Above 1.25 a glass is
    called spectrally selective: it admits daylight while rejecting heat, which is the whole
    argument for solar-control glazing in Hyderabad and Vijayawada.
    """
    vlt = _num(vals.get("visible_light_transmission_pct"))
    shgc = _num(vals.get("solar_heat_gain_coefficient"))
    if not vlt or not shgc:
        return None
    return round((vlt / 100) / shgc, 2), "LSG = visible light transmission / solar heat gain coefficient; above 1.25 is spectrally selective"


def derive_shading_coefficient(vals: dict) -> tuple[float, str] | None:
    """SC = SHGC / 0.87, the older figure Indian tender documents still specify against."""
    shgc = _num(vals.get("solar_heat_gain_coefficient"))
    if not shgc:
        return None
    return round(shgc / 0.87, 2), "SC = SHGC / 0.87, referenced to 3 mm clear float glass"


def derive_fill_factor(vals: dict) -> tuple[float, str] | None:
    """
    FF = (Vmp x Imp) / (Voc x Isc). How square the I-V curve is, and the single number that
    separates a well-made cell from a poorly made one at the same rated wattage.
    """
    vmp, imp = _num(vals.get("vmp_v")), _num(vals.get("imp_a"))
    voc, isc = _num(vals.get("voc_v")), _num(vals.get("isc_a"))
    if not all((vmp, imp, voc, isc)) or voc * isc == 0:
        return None
    return round((vmp * imp) / (voc * isc), 3), "FF = (Vmp × Imp) / (Voc × Isc), how square the I-V curve is"


DERIVED = {
    "bulbs": [
        {
            "key": "luminous_intensity",
            "group": "physical",
            "label": "Luminous intensity",
            "unit": "cd",
            "data_type": "number",
            "fn": derive_luminous_intensity,
            "importance_rank": 1,
            "compare": True,
        }
    ],
    "glass": [
        {
            "key": "light_to_solar_gain",
            "group": "physical",
            "label": "Light-to-solar-gain ratio",
            "unit": None,
            "data_type": "number",
            "fn": derive_light_to_solar_gain,
            "importance_rank": 1,
            "compare": True,
        },
        {
            "key": "shading_coefficient",
            "group": "physical",
            "label": "Shading coefficient",
            "unit": None,
            "data_type": "number",
            "fn": derive_shading_coefficient,
            "importance_rank": 2,
            "compare": True,
        },
    ],
    "solar-panels": [
        {
            "key": "fill_factor",
            "group": "physical",
            "label": "Fill factor",
            "unit": None,
            "data_type": "number",
            "fn": derive_fill_factor,
            "importance_rank": 2,
            "compare": True,
        }
    ],
}

# ── workbook styling ────────────────────────────────────────────────────────────
TEAL = "087F80"
INK = "0F1B1C"
BAND = "F1F6F6"
RULE = "D7E2E2"
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FONT = Font(bold=True, color=INK, size=10)
BODY_FONT = Font(size=10)
MONO_FONT = Font(size=9, color="55696B")
THIN = Side(style="thin", color=RULE)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Group", 26),
    ("Specification", 34),
    ("Key", 30),
    ("Type", 9),
    ("Unit", 10),
    ("Filter", 8),
    ("Key spec", 9),
    ("Card", 7),
    ("Compare", 9),
    ("Rank", 6),
]
PER_SKU = [("{code}", 42), ("Source", 11)]
TRAILING = [("Notes", 46)]


def clean(text: str) -> str:
    """Undo the encoding damage in the old workbook and normalise whitespace."""
    return re.sub(r"\s+", " ", text.replace("�", "—")).strip()


def is_placeholder(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in PLACEHOLDERS


def facet_keys(slug: str) -> set[str]:
    """
    Which attributes are facets is the filter policy's decision, not a second opinion held in
    the workbook. Reading it here means the Filter column and registry/filter-policy.json
    cannot disagree — they did, and the policy named eight attributes the registry marked
    unfilterable or had dropped outright.
    """
    policy = json.loads((REGISTRY / "filter-policy.json").read_text(encoding="utf-8"))["categories"].get(slug, {})
    keys: set[str] = set()
    for section in ("primary", "more", "toolbar"):
        keys.update(policy.get(section, []))
    keys.update(policy.get("conditional", {}))
    if policy.get("lead"):
        keys.add(policy["lead"])
    return keys - {"price", "brand", "stock"}  # the engine computes these itself


def load_category(slug: str) -> dict:
    registry = json.loads((REGISTRY / f"{slug}.json").read_text(encoding="utf-8"))
    skus = []
    for path in sorted((CURATED / slug).glob("*.json")):
        sku = json.loads(path.read_text(encoding="utf-8"))
        if sku["sku_code"] not in EXCLUDE_SKUS:
            skus.append(sku)
    return {"registry": registry, "skus": skus}


def build_rows(slug: str, registry: dict, skus: list[dict]) -> list[dict]:
    """One row per attribute that at least one SKU has a real value for."""
    defs = {a["key"]: a for a in registry["attributes"]}
    facets = facet_keys(slug)

    # Plain {key: value} per SKU, used by the derivations.
    plain = [
        {k: v.get("value") for k, v in s["attributes"].items() if not is_placeholder(v.get("value"))}
        for s in skus
    ]

    rows: list[dict] = []
    for key, a in defs.items():
        if JUNK_KEY.search(key):
            continue
        cells = []
        for s in skus:
            raw = s["attributes"].get(key)
            if not raw or is_placeholder(raw.get("value")):
                cells.append(None)
            else:
                cells.append(
                    {
                        "value": raw["value"],
                        "provenance": raw.get("provenance", "ai_filled"),
                        "note": clean(str(raw.get("note") or "")),
                    }
                )
        if not any(cells):
            continue  # defined but never populated: it cannot render, so it is not a row
        rows.append(
            {
                "group": a["group"],
                "label": clean(a["label"]),
                "key": key,
                "data_type": a["data_type"],
                "unit": a.get("unit"),
                "is_filterable": key in facets,
                "importance_rank": a.get("importance_rank", 3),
                "compare": a.get("compare", False),
                "cells": cells,
            }
        )

    for spec in DERIVED.get(slug, []):
        cells, notes = [], ""
        for vals in plain:
            got = spec["fn"](vals)
            if got is None:
                cells.append(None)
            else:
                value, formula = got
                notes = notes or formula
                cells.append({"value": value, "provenance": "derived", "note": formula})
        if not any(cells):
            continue
        rows = [r for r in rows if r["key"] != spec["key"]]  # replace any prose stand-in
        rows.append(
            {
                "group": spec["group"],
                "label": spec["label"],
                "key": spec["key"],
                "data_type": spec["data_type"],
                "unit": spec["unit"],
                "is_filterable": spec["key"] in facets,
                "importance_rank": spec["importance_rank"],
                "compare": spec["compare"],
                "cells": cells,
            }
        )
    return rows


def regroup(rows: list[dict], slug: str, spec_groups: dict) -> list[dict]:
    """
    Re-file every row under the heading spec-groups.json puts it in, and order it by the
    position it holds there. The registry's own group came from whichever of the old
    template's 24 columns the label happened to sit in, which is arbitrary; this is not.
    Any key not listed for the category is reported so the map can never silently drift.
    """
    plan = spec_groups["categories"][slug]
    placement = {key: (gi, ki, group) for gi, (group, keys) in enumerate(plan.items()) for ki, key in enumerate(keys)}

    unplaced = [r["key"] for r in rows if r["key"] not in placement]
    if unplaced:
        raise SystemExit(f"spec-groups.json is missing {slug} keys: {unplaced}")
    listed = {k for keys in plan.values() for k in keys}
    stale = sorted(listed - {r["key"] for r in rows})

    for r in rows:
        gi, ki, group = placement[r["key"]]
        r["group"], r["_gi"], r["_ki"] = group, gi, ki
    if stale:
        print(f"    note: {slug} lists {len(stale)} key(s) with no value in any SKU: {', '.join(stale)}")
    return sorted(rows, key=lambda r: (r["_gi"], r["_ki"]))


def write_sheet(wb: Workbook, slug: str, meta: dict, rows: list[dict], skus: list[dict], labels: dict) -> None:
    keyset, cardset = set(KEY_SPECS.get(slug, [])), set(CARD_SPECS.get(slug, []))
    ws = wb.create_sheet(meta["name"][:31])

    header = [c[0] for c in COLUMNS]
    widths = [c[1] for c in COLUMNS]
    for s in skus:
        header += [s["sku_code"], "Source"]
        widths += [w for _, w in PER_SKU]
    header += [c[0] for c in TRAILING]
    widths += [c[1] for c in TRAILING]

    filled = sum(1 for r in rows for c in r["cells"] if c)
    ws.cell(row=1, column=1).value = (
        f"{meta['name'].upper()}  —  {DEPARTMENT[slug]}  ·  "
        f"{len(skus)} brands  ·  {len(rows)} specifications  ·  {filled} values  ·  "
        f"unit: {meta['unit']}  ·  GST {meta['gst_rate']:g}%"
    )
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color=TEAL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    ws.row_dimensions[1].height = 24

    for i, (name, width) in enumerate(zip(header, widths), start=1):
        c = ws.cell(row=2, column=i)
        c.value = name
        c.font = HEAD_FONT
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 30

    r = 3
    last_group = None
    for row in rows:
        banded = labels.get(row["group"], row["group"]) != last_group
        last_group = labels.get(row["group"], row["group"])
        values = [
            last_group if banded else "",
            row["label"],
            row["key"],
            row["data_type"],
            row["unit"] or "",
            "yes" if row["is_filterable"] else "",
            "yes" if row["key"] in keyset else "",
            "yes" if row["key"] in cardset else "",
            "yes" if row["compare"] else "",
            row["importance_rank"],
        ]
        note = ""
        for cell in row["cells"]:
            values += ["" if not cell else cell["value"], "" if not cell else cell["provenance"]]
            if cell and cell["note"]:
                note = note or cell["note"]
        values.append(note)

        for i, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=i)
            c.value = v
            c.border = CELL_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=i in (2, 11, 13, 15, len(header)))
            if i == 1:
                c.font = GROUP_FONT
            elif i in (3, len(header)):
                c.font = MONO_FONT
            else:
                c.font = BODY_FONT
            if row["key"] in keyset and i == 2:
                c.fill = PatternFill("solid", fgColor=BAND)
        r += 1

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(header))}{r - 1}"


def write_overlay(cats: list[dict]) -> int:
    """
    Search and faceting metadata the sheet deliberately does not carry: which widget a
    filter draws as, the order filters sit in, an enum's permitted values, and the synonyms
    Meilisearch expands a query with. None of it is a product specification and none of it
    belongs in a document a product manager edits, but all of it is real work that would be
    lost if the registry were regenerated from the sheet alone. Captured here from the
    registries as they stand and merged back in by from-sheet.ts.
    """
    target = REGISTRY / "attribute-overlay.json"
    if target.exists():
        return -1  # a one-time capture; re-running must not narrow it to today's attribute set
    overlay: dict[str, dict] = {}
    kept = 0
    for meta in cats:
        slug = meta["slug"]
        entries = {}
        for a in json.loads((REGISTRY / f"{slug}.json").read_text(encoding="utf-8"))["attributes"]:
            e = {}
            if a.get("filter_widget"):
                e["filter_widget"] = a["filter_widget"]
            if a.get("filter_order", 100) != 100:
                e["filter_order"] = a["filter_order"]
            if a.get("enum_values"):
                e["enum_values"] = a["enum_values"]
            if a.get("synonyms"):
                e["synonyms"] = a["synonyms"]
            if e:
                entries[a["key"]] = e
                kept += 1
        overlay[slug] = entries
    (REGISTRY / "attribute-overlay.json").write_text(
        json.dumps(
            {
                "$comment": (
                    "Faceting and search metadata per category, merged onto the registry that "
                    "from-sheet.ts builds from WHOLE_PRODUCT_LIST_BO_PRODUCT_CALENDAR.xlsx. The "
                    "workbook owns what a product IS; this owns how it is filtered and found. "
                    "Generated once by tools/build_calendar_workbook.py; edit by hand thereafter."
                ),
                "categories": overlay,
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    return kept


def main() -> None:
    spec_groups = json.loads((REGISTRY / "spec-groups.json").read_text(encoding="utf-8"))
    cats = json.loads((REGISTRY / "categories.json").read_text(encoding="utf-8"))["categories"]
    labels = spec_groups["labels"]

    wb = Workbook()
    wb.remove(wb.active)

    total_rows = total_values = 0
    for meta in sorted(cats, key=lambda c: c["display_order"]):
        slug = meta["slug"]
        data = load_category(slug)
        rows = regroup(build_rows(slug, data["registry"], data["skus"]), slug, spec_groups)
        write_sheet(wb, slug, meta, rows, data["skus"], labels)
        values = sum(1 for r in rows for c in r["cells"] if c)
        total_rows += len(rows)
        total_values += values
        missing = [k for k in KEY_SPECS.get(slug, []) if k not in {r["key"] for r in rows}]
        flag = f"   MISSING KEY SPECS: {missing}" if missing else ""
        print(f"  {meta['name']:<20} {len(rows):>4} specs  {values:>5} values  {len(data['skus'])} brands{flag}")

    wb.save(OUT)
    kept = write_overlay(cats)
    print(f"\n{len(cats)} sheets, {total_rows} specifications, {total_values} values -> {OUT.name}")
    if kept >= 0:
        print(f"{kept} attributes carry facet/search metadata -> registry/attribute-overlay.json")
    else:
        print("registry/attribute-overlay.json already exists and was left as it is")


if __name__ == "__main__":
    main()
